import pandas as pd
import openpyxl
import re
import streamlit as st
from datetime import datetime

def atualizar_medicoes(arquivo_valores, arquivo_acompanhamento):
    """
    Atualiza a planilha de Acompanhamento de Medições com dados de um arquivo de valores.
    """
    try:
        # Carregar os arquivos Excel
        df_valores = pd.read_excel(arquivo_valores)

        # Tratar colunas de data
        df_valores['DataIni_cvm'] = pd.to_datetime(df_valores['DataIni_cvm'], errors='coerce')
        df_valores['DataEmis_Nf'] = pd.to_datetime(df_valores['DataEmis_Nf'], errors='coerce')
        df_valores['DataRec'] = pd.to_datetime(df_valores['DataRec'], errors='coerce')
        
        df_valores.dropna(subset=['DataIni_cvm'], inplace=True)

        # Mapear o mês e ano para o formato desejado
        meses_map = {
            1: 'JANEIRO', 2: 'FEVEREIRO', 3: 'MARÇO', 4: 'ABRIL',
            5: 'MAIO', 6: 'JUNHO', 7: 'JULHO', 8: 'AGOSTO',
            9: 'SETEMBRO', 10: 'OUTUBRO', 11: 'NOVEMBRO', 12: 'DEZEMBRO'
        }
        df_valores['MesAno'] = df_valores['DataIni_cvm'].apply(
            lambda x: f"{meses_map[x.month]}/{str(x.year)[2:]}"
        )

        # Agrupar os dados
        agg_funcs = {
            'TotalValorNF': 'sum',
            'DataEmis_Nf': 'max',
            'DataRec': 'max'
        }
        df_aggregated = df_valores.groupby(['ObraVen_Vnv', 'MesAno']).agg(agg_funcs).reset_index()

        # Carregar a planilha de acompanhamento
        workbook = openpyxl.load_workbook(arquivo_acompanhamento)
        
        st.write("Planilha de acompanhamento carregada. Iniciando o processo de atualização...")

        # Iterar sobre cada linha da planilha de valores agrupados
        for _, row in df_aggregated.iterrows():
            obra_id = str(int(row['ObraVen_Vnv']))
            mes_ano_para_encontrar = row['MesAno'].upper()
            
            aba_encontrada = False
            # Procurar a obra na planilha de Acompanhamento de Medições
            for sheet_name in workbook.sheetnames:
                match = re.match(r'^(\d+)', sheet_name)
                if match and match.group(1) == obra_id:
                    sheet = workbook[sheet_name]
                    aba_encontrada = True
                    
                    header_map = {}
                    header_row_index = -1
                    
                    # Procurar o cabeçalho na aba
                    for i, row_sheet in enumerate(sheet.iter_rows(min_row=1, max_row=10)):
                        for j, cell in enumerate(row_sheet):
                            if cell.value and isinstance(cell.value, str):
                                if "MÊS" in cell.value.upper(): header_map['MÊS'] = j + 1
                                if "DATA NF" in cell.value.upper(): header_map['DATA NF'] = j + 1
                                if "VALOR NF" in cell.value.upper(): header_map['VALOR NF'] = j + 1
                                if "DATA DO RECEBIMENTO" in cell.value.upper(): header_map['DATA DO RECEBIMENTO'] = j + 1
                        if len(header_map) >= 4:
                            header_row_index = i + 1
                            break
                    
                    if header_row_index == -1:
                        st.warning(f"Aviso: Não foi possível encontrar o cabeçalho na aba '{sheet_name}'. Pulando esta aba.")
                        continue

                    for i in range(header_row_index + 1, sheet.max_row + 1):
                        mes_na_planilha = sheet.cell(row=i, column=header_map['MÊS']).value
                        
                        if mes_na_planilha and isinstance(mes_na_planilha, str) and mes_na_planilha.upper() == mes_ano_para_encontrar:
                            st.success(f"Correspondência encontrada! Obra: {obra_id}, Mês: {mes_ano_para_encontrar}, Aba: '{sheet_name}'")
                            
                            # Preencher colunas
                            if 'DATA NF' in header_map and not pd.isna(row['DataEmis_Nf']):
                                sheet.cell(row=i, column=header_map['DATA NF']).value = row['DataEmis_Nf'].date()
                            if 'VALOR NF' in header_map:
                                sheet.cell(row=i, column=header_map['VALOR NF']).value = row['TotalValorNF']
                            if 'DATA DO RECEBIMENTO' in header_map and not pd.isna(row['DataRec']):
                                sheet.cell(row=i, column=header_map['DATA DO RECEBIMENTO']).value = row['DataRec'].date()
                            
                            break
            
            if not aba_encontrada:
                st.warning(f"Aviso: Obra com ID '{obra_id}' não encontrada nas abas da planilha de acompanhamento.")

        # Salvar o arquivo atualizado em memória
        from io import BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    except Exception as e:
        st.error(f"Ocorreu um erro inesperado: {e}")
        return None

# --- INÍCIO DA EXECUÇÃO COM STREAMLIT ---

st.title("Atualização de Acompanhamento de Medições")

# Seleção de arquivos com Streamlit
arquivo_valores = st.file_uploader("Escolha a planilha de VALORES (Pasta1.xlsx)", type=["xlsx"])
arquivo_acompanhamento = st.file_uploader("Escolha a planilha de ACOMPANHAMENTO DE MEDIÇÕES", type=["xlsx"])

if arquivo_valores and arquivo_acompanhamento:
    if st.button("Atualizar Medições"):
        with st.spinner("Atualizando medições..."):
            resultado = atualizar_medicoes(arquivo_valores, arquivo_acompanhamento)
            
            if resultado:
                st.success("Processo concluído com sucesso!")
                st.download_button(
                    label="Baixar Arquivo Atualizado",
                    data=resultado,
                    file_name="ACOMPANHAMENTO_MEDICOES_ATUALIZADO.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
# python -m streamlit run planilhaAlvaro.py
